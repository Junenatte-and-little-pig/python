# -*- encoding: utf-8 -*-
from openpyxl import *

wb = load_workbook('person1.xlsx')
print(wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(dir(type(ws)))
    # print(ws.cell(row, column))
    for c in ws.columns:
        # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.A2>)
        print(c)  # <class 'tuple'>
        for cell in c:
            print(cell)  # <Cell 'Sheet'.A1>
            print(
                'value:{0}{1},{0}column:{0}{2},{0}column_letter:{0}{3},{0}row:{0}{4}'.format(
                    '\t', cell.value, cell.column, cell.column_letter,
                    cell.row))
    for r in ws.rows:
        # (< Cell 'Sheet'.A1 >,)
        # (< Cell 'Sheet'.A2 >,)
        print(r)  # <class 'tuple'>

    print(ws['A1'])
    print(ws['A1:B2'])
    for row in ws['A1:A2']:
        # 指定区域内行元组组成的元组
        for cell in row:
            print(
                'value:{0}{1},{0}column:{0}{2},{0}column_letter:{0}{3},{0}row:{0}{4}'.format(
                    '\t', cell.value, cell.column, cell.column_letter,
                    cell.row))
    for row in ws:
        # 全工作表内行元组组成的元组
        for cell in row:
            print(
                'value:{0}{1},{0}column:{0}{2},{0}column_letter:{0}{3},{0}row:{0}{4}'.format(
                    '\t', cell.value, cell.column, cell.column_letter,
                    cell.row))
print(wb.worksheets)
for work_sheet in wb.worksheets:  # same as wb[sheet_name]
    print(work_sheet.title)  # same as sheet_name
