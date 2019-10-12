# -*- encoding: utf-8 -*-
from openpyxl import *

wb=Workbook()
ws=wb.active
ws['A1']='zhangsan'
ws['B1']=18
ws['A2']='lisi'
ws['B2']=20
wb.save('person1.xlsx')